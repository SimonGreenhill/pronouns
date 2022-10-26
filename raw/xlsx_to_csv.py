#!/usr/bin/env python3
# coding=utf-8
"""..."""
__author__ = 'Simon J. Greenhill <simon@simon.net.nz>'
__copyright__ = 'Copyright (c) 2022 Simon J. Greenhill'
__license__ = 'New-style BSD'

import csvw
from openpyxl import load_workbook

EXPECTED = [ 
    'word', 'ipa', 'parameter', 'description', 'alternative', 'comment',
    'translation', 'glottocode', 'notes'
]

OUTHEADER = "word,ipa,parameter,description,localid,alternative,comment,translation,glottocode,source".split(",")

def read(filename):
    wb = load_workbook(filename)
    header = None
    for i, row in enumerate(wb.active.iter_rows(values_only=True)):
        if i == 0:
            header = row
            continue
        
        row = dict(zip(header, row))
        # check expected
        missing = [e for e in EXPECTED if e not in row]
        if len(missing):
            raise ValueError("Missing column %s" % e)
        # move notes -> source
        row['source'] = row['notes']
        del(row['notes'])
        
        yield row


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Does something.')
    parser.add_argument("xlsx", help='xlsx filename (INPUT)')
    parser.add_argument("csv", help='csv filename (OUTPUT)')
    args = parser.parse_args()

    with csvw.UnicodeWriter(args.csv) as writer:
        writer.writerow(OUTHEADER)
        for row in read(args.xlsx):
            print(row)
            writer.writerow([row.get(h, "") for h in OUTHEADER])

